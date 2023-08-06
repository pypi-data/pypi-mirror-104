from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal

from ofxstatement.parser import StatementParser
from ofxstatement.plugin import Plugin
from ofxstatement import statement

# file format options
t_time_format = '%d/%m/%Y'
t_encoding = 'utf8'

def find_transaction_list_start_row_idx(worksheet):
    for i in range(1, worksheet.max_row):
        cell = worksheet.cell(row=i, column=1)
        if cell.value is not None and cell.value == u'Ngày giao dịch':
            return i + 1
    pass

def find_transaction_list_end_row_idx(worksheet):
    for i in range(1, worksheet.max_row):
        cell = worksheet.cell(row=i, column=1)
        if cell.value is not None and cell.value == u'Tổng':
            return i - 1
    pass

class VietcombankExcelStatementParser(StatementParser):
    statement = None

    def __init__(self, fin):
        super().__init__()
        self.statement = statement.Statement()

        wb = load_workbook(fin)
        self.worksheet = wb.worksheets[0]
        self.start_row = find_transaction_list_start_row_idx(self.worksheet)
        self.end_row = find_transaction_list_end_row_idx(self.worksheet)

    def split_records(self):
        return range(self.start_row, self.end_row + 1)

    def parse_record(self, line):
        transaction = statement.StatementLine()

        transaction.date = datetime.strptime(self.worksheet.cell(row=line, column=1).value, t_time_format)

        transaction.id = self.worksheet.cell(row=line, column=2).value

        transaction.amount = Decimal(self.worksheet.cell(row=line, column=4).value.replace(',', ''))

        direction = self.worksheet.cell(row=line, column=3).value
        if direction == '-':
            transaction.trntype = 'CREDIT'
            transaction.amount *= -1
        else:
            transaction.trntype = 'DEBIT'

        transaction.memo = self.worksheet.cell(row=line, column=5).value
        return transaction

class VietcombankExcelPlugin(Plugin):
    """Vietcombank
    """

    def get_parser(self, fin):
        if not fin.endswith('.xlsx'):
            raise f'Invalid report file {fin}. Expected .xslx'

        parser = VietcombankExcelStatementParser(fin)
        parser.statement.currency = self.settings['currency']
        parser.statement.account_id = self.settings['account']
        parser.statement.bank_id = self.settings.get('bank', 'Vietcombank')
        return parser
